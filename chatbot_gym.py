import pandas as pd
import os
import json
import torch
import torch.cuda
from tqdm import tqdm
import numpy as np
import pickle as pkl
import regex as re
import random as rd
from openpyxl import load_workbook
from sklearn.feature_extraction.text import TfidfVectorizer
from dotenv import load_dotenv


load_dotenv()

FILE_PATH = os.getenv("FILE_PATH")
GYM_VOCAB_FILE = os.getenv("GYM_VOCAB_FILE")
PATH_DATASET1 = os.getenv("PATH_DATASET1")
PATH_DATASET2 = os.getenv("PATH_DATASET2")
GYM_TOKENIZER = os.getenv("GYM_TOKENIZER")
GYM_PATH_MODEL = os.getenv("GYM_PATH_MODEL")



def windows(tokenized_list):
    text_tokenized_window = []
    start_index = 0
    windows_size = 10
    end_index = len(tokenized_list) - windows_size

    while start_index < end_index:
        text_tokenized_window.append(tokenized_list[start_index:start_index+windows_size+1])
        start_index += 1

    return text_tokenized_window

def excel_results_registration(new_data, prediccion=False):
        """
        Añade registros al archivo Excel debajo de la última línea.
        
        :param file_path: Ruta del archivo Excel.
        :param sheet_name: Nombre de la hoja donde se agregarán los registros.
        :param new_data: Lista de listas con los datos a añadir.
        """
        file_path = FILE_PATH
        sheet_name = "Registro_ficheros"
        try:
            # Cargar el archivo Excel
            workbook = load_workbook(file_path)
            
            # Seleccionar la hoja
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"La hoja '{sheet_name}' no existe en el archivo Excel.")
            
            if prediccion == False:
            
                sheet = workbook[sheet_name]

                # Encontrar la última fila con datos
                last_row = sheet.max_row + 1

                # Añadir los nuevos datos debajo de la última fila
                sheet.cell(row=last_row, column=1, value=new_data)

                # Guardar los cambios
                workbook.save(file_path)
                print(f"Datos añadidos correctamente al archivo '{file_path}'.")

            else:
                sheet = workbook[sheet_name]

                # Encontrar la última fila con datos
                last_row = sheet.max_row 

                # Añadir los nuevos datos debajo de la última fila
                sheet.cell(row=last_row, column=2, value=new_data)

                # Guardar los cambios
                workbook.save(file_path)
                print(f"Datos añadidos correctamente al archivo '{file_path}'.")

        except Exception as e:
            print(f"Error al añadir datos al archivo Excel: {e}")



def dataset_to_text_list(dataset):
    text_list = ""
    for line in dataset:
        line = line.replace("\n", "").rstrip()
        if "-" in line:
            line = line.replace("-", "")
            text_list += line
        else:
            text_list += line+" "
    return text_list


class Tokenizer():
    def __init__(self):
        pass
    def create_vocabulary(self, text_list):
        if os.path.exists(GYM_VOCAB_FILE) == False:

            VOCABULARY = {}
            print(f"Entrando en el proceso de la creacion del vocabulario con un total de {len(text_list.split())} palabras")

            text_list_splitted = text_list.split()
            number = 0

            for word in tqdm(text_list_splitted, desc="Progreso"):
                    if word not in VOCABULARY:
                        chequeos = True
                        if re.findall("[\¿\!\#\$\%\&\'\(\)\*\+\,\-\.\/\:\;\<\=\>\?\@\[\]\^\_\`\{\|\}\~\“]+", word):   #Tengo que meter todos los if necesarios para solo eliminar palabras que no me interesen rollo numeros, simbolos
                            chequeos = False
                        elif re.findall("^(?=.*[A-Za-z])(?=.*\d)[A-Za-z\d]+$", word):
                            chequeos = False   
                        elif re.findall("^[0-9]+$", word):
                            chequeos = False 
                        elif re.findall("[∀∂∃∅∆∇∈∉∋∏∑−∗√∝∞∠∧∨∩∪∫∴∼≃≅≈≠≡≤≥⊂⊃⊄⊆⊇⊕⊗⊥⋅⌈⌉⌊⌋⟨⟩◊♠♣♥♦♪♭♯✓✔✕✖✗✘✦★☆✩✪✮✯✰✳✴✵✶✷✸✹✺✻✼✽✾✿❀❁❂❃❄❅❆❇❈❉❊❋❍❏❐❑❒❖❦❧➔➡➞➟➠➢➣➤➥➦➧➨➩➪➫➬➭➮➯➱➲➳➴➵➸➹➺➻➼➽➾⟰⟱⟲⟳⟴⟵⟶⟷⟸⟹⟺⟻⟼⟽⟾⟿⦀⦁⦂⦃⦄⦅⦆⦇⦈⦉⦊⦋⦌⦍⦎⦏⦐⦑⦒⦓⦔⦕⦖⦗⦘⧾⧿⩀⩁⩂⩃⩄⩅⩆⩇⩈⩉⩊⩋⩌⩍⩎⩏⩐⩑⩒⩓⩔⩕⩖⩗⩘⩙⩚⩛⩜⩝⩞⩟⩠⩡⩢⩣⩤⩥⩦⩧⩨⩩⩪⩫⩬⩭⩮⩯⩰⩱⩲⩳⩴⩵⩶⩷⩸⩹⩺⩻⩼⪽⪾⪿⫀⫁⫂⫃⫄⫅⫆⫇⫈⫉⫊⫋⫌⫍⫎⫏⫐⫑⫒⫓⫔⫕⫖⫗⫘⫙⫚⫛⫝̸⫝⫞⫟⫠⫡⫢⫣⫤⫥⫦⫧⫨⫩⫪⫫⫬⫭⫮⫯⫰⫱⫲⫳⫴⫵⫶⫷]+", word):
                            chequeos = False
                        elif re.findall("[¡¢£¤¥¦§¨©ª«¬­®¯°±²³´µ¶·¸¹º»¼½¾¿×÷ˆ˜‐‑‒–—―‘’‚‛“”„†‡•…‰′″‹›‾⁄€™←↑→↓↔↕↨⇄⇅⇆⇒⇔]+",word):
                            chequeos = False

                        if chequeos == True: 
                            number += 1            
                            VOCABULARY[word] = number

                        
            with open(GYM_VOCAB_FILE, "wb") as gv:
                pkl.dump(VOCABULARY, gv)
                print(f"Vocabulario guardado con {len(VOCABULARY)}")
        else:
            with open(GYM_VOCAB_FILE, "rb") as ml:
                VOCABULARY = pkl.load(ml)

            print("Ya hay un vocabulario entrenado")

        # Update the class attribute
        self.VOCABULARY = VOCABULARY
        return self.VOCABULARY

    def text_to_seq(self, string):
        seq = []
        for word in string.split():
            if word in self.VOCABULARY:
                value = self.VOCABULARY[word]
                seq.append(value)
            else: 
                continue

        return seq
    
    def seq_to_text(self, string_tokened):
        text_destokenized = []
        for word in string_tokened:
            word_decoded = list(self.VOCABULARY.keys())[list(self.VOCABULARY.values()).index(word)]
            text_destokenized.append(word_decoded)
        return " ".join(text_destokenized)



    
class CharRNNDataset(torch.utils.data.Dataset):
    def __init__(self, text_encoded_windows, train=True):
        self.text = text_encoded_windows
        self.train = train

    def __len__(self):
        return len(self.text)

    def __getitem__(self, ix):
        if self.train:
            return torch.tensor(self.text[ix][:-1]), torch.tensor(self.text[ix][-1])
        return torch.tensor(self.text[ix])


class CharRNN(torch.nn.Module):
    def __init__(self, input_size, embedding_size=256, hidden_size=512, num_layers=2, dropout=0.2):
        self.text_output = f"Parametros para este modelo: embedding_size = {embedding_size}, hidden_size = {hidden_size}, num_layers = {num_layers}, dropout = {dropout}"
        super().__init__()
        self.encoder = torch.nn.Embedding(input_size, embedding_size)
        self.rnn = torch.nn.LSTM(input_size=embedding_size, hidden_size=hidden_size, num_layers=num_layers, dropout=dropout, batch_first=True)
        self.fc = torch.nn.Linear(hidden_size, input_size)

    def forward(self, x):
        x = self.encoder(x)
        x, h = self.rnn(x)      
        y = self.fc(x[:,-1,:])
        return y

    def get_output_model(self):
        return self.text_output 
  
device = "cuda" if torch.cuda.is_available() else "cpu"

def fit(model, dataloader, epochs=10):
    output = ""
    model.to(device)
    optimizer = torch.optim.Adam(model.parameters(), lr=1e-2)
    criterion = torch.nn.CrossEntropyLoss()
    for epoch in range(1, epochs+1):
        model.train()
        train_loss = []
        bar = tqdm(dataloader['train'])
        for batch in bar:
            X, y = batch
            X, y = X.to(device), y.to(device)
            optimizer.zero_grad()
            y_hat = model(X)
            loss = criterion(y_hat, y)
            loss.backward()
            optimizer.step()
            train_loss.append(loss.item())
            bar.set_description(f"loss {np.mean(train_loss):.5f}")
        bar = tqdm(dataloader['val'])
        val_loss = []
        with torch.no_grad():
            for batch in bar:
                X, y = batch
                X, y = X.to(device), y.to(device)
                y_hat = model(X)
                loss = criterion(y_hat, y)
                val_loss.append(loss.item())
                bar.set_description(f"val_loss {np.mean(val_loss):.5f}")
        print(f"Epoch {epoch}/{epochs} loss {np.mean(train_loss):.5f} val_loss {np.mean(val_loss):.5f}")
        output += f"/////---{epoch} -> train_loss = {np.mean(train_loss):.5f}, val_loss -> {np.mean(val_loss):.5f} "
    
    return output

def predict(model, X):
    model.eval() 
    with torch.no_grad():
        X = torch.tensor(X).long().to(device)
        pred = model(X.unsqueeze(0))
        return pred

def model_train():
    # Entrenar el modelo con mas textos
    output = ""

    with open(PATH_DATASET1 ,errors="ignore", encoding="utf8") as dataset:
        dataset = dataset.readlines()
        text_list1 = dataset_to_text_list(dataset=dataset)
    with open(PATH_DATASET2,errors="ignore", encoding="utf8") as dataset:
        dataset = dataset.readlines()
        text_list2 = dataset_to_text_list(dataset=dataset)
    text_list = text_list1 + " " + text_list2

    
    print(text_list[-10:])

    print("Datasets de textos leido, procedemos a su estudio")

    tokenizer = Tokenizer()
    VOCABULARY = tokenizer.create_vocabulary(text_list=text_list)
    text_tokenized = tokenizer.text_to_seq(text_list)
    train_size = len(text_tokenized) * 80 // 100
    train = text_tokenized[:train_size]
    test = text_tokenized[train_size:]

    train_text_encoded_windows = windows(train)
    print(train_text_encoded_windows[0:10])
    print(f"Se esta entrenando con {len(train_text_encoded_windows)} ventanas")
    test_text_encoded_windows = windows(test)
    print(test_text_encoded_windows[0:10])
    print(f"Se esta evaluando con {len(test_text_encoded_windows)} ventanas")

    dataset = {
        'train': CharRNNDataset(train_text_encoded_windows),
        'val': CharRNNDataset(test_text_encoded_windows)
    }

    dataloader = {
        'train': torch.utils.data.DataLoader(dataset['train'], batch_size=128, shuffle=True, pin_memory=True),
        'val': torch.utils.data.DataLoader(dataset['val'], batch_size=512, shuffle=False, pin_memory=True),
    }
    

    model = CharRNN(input_size=max(VOCABULARY.values())+1)
    output_inicio = model.get_output_model()
    output = fit(model, dataloader, epochs=3)
    output_final = output_inicio + output

    excel_results_registration(output_final)

    print(model_serialize(model))

    with open(GYM_TOKENIZER, "wb") as tf:
        pkl.dump(tokenizer, tf)

    return r"The model was train"

def model_serialize(model):
    with open(GYM_PATH_MODEL, "wb") as mf:
        pkl.dump(model, mf)
    return f"The model was serialize in {os.getcwd()}"

def model_load():
    with open(GYM_PATH_MODEL, "rb") as ml:
        model = pkl.load(ml)
    with open(GYM_TOKENIZER, "rb") as tl:
        tokenizer = pkl.load(tl)
    with open(GYM_VOCAB_FILE, "rb") as vj:
        vocabulary = pkl.load(vj)

    return model, tokenizer, vocabulary

def model_LLM_response(input):
    model, tokenizer, vocabulary = model_load()

    temp=0.8
    for i in range(30):
        X_new_encoded = tokenizer.text_to_seq(input)
        y_pred = predict(model, X_new_encoded)
        y_pred = y_pred.view(-1).div(temp).exp()
        top_i = torch.multinomial(y_pred, 1)[0]
        predicted_char = list(vocabulary.keys())[list(vocabulary.values()).index(int(top_i))]
        input += " "+predicted_char

    return input 

"""
if __name__ == "__main__":


    #model_train()
    model, tokenizer, vocabulary = model_load()

    X_new = "Los abdominales son"
    #print(tokenizer.text_to_seq(X_new))

    temp=0.8
    for i in range(30):
        X_new_encoded = tokenizer.text_to_seq(X_new)
        y_pred = predict(model, X_new_encoded)
        y_pred = y_pred.view(-1).div(temp).exp()
        top_i = torch.multinomial(y_pred, 1)[0]
        predicted_char = list(vocabulary.keys())[list(vocabulary.values()).index(int(top_i))]
        X_new += " "+predicted_char

    excel_results_registration(X_new, prediccion=True)
    print(X_new)


    for i in range(30):
    #Se puede añadir un parametros de temperatura para elegir entre las palabras con mayor probabildad aleatoriamente
        X_new_encoded = tokenizer.text_to_seq(X_new)

        y_pred = predict(model, X_new_encoded)
        y_pred = torch.argmax(y_pred, axis=1)[0].item()

        X_new += " "+tokenizer.seq_to_text([y_pred])


"""

       

